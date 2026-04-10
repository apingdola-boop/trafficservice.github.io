"""
엑셀 기반 Gmail SMTP 발송 — 레거시 실습 스크립트
비밀번호는 환경 변수로만 설정하세요.
"""
import os
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook

EXCEL_FILE = os.environ.get("EXCEL_FILE", "sample.xlsx")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "")
RECEIVER_EMAIL = os.environ.get("RECEIVER_EMAIL", "")
MAX_EMAILS = int(os.environ.get("MAX_EMAILS", "10"))

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

GMAIL_ADDRESS = os.environ.get("GMAIL_ADDRESS", "").strip()
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "").strip()


def read_excel_data(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    print(f"엑셀 파일 컬럼: {headers}")

    customers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            customers.append(dict(zip(headers, row)))

    return headers, customers


def create_email_message(sender, receiver, subject, body):
    message = MIMEMultipart()
    message["From"] = sender
    message["To"] = receiver
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain", "utf-8"))
    return message


def send_email(smtp_connection, sender, receiver, message):
    smtp_connection.sendmail(sender, receiver, message.as_string())


def main():
    print("=" * 60)
    print("엑셀 → Gmail 발송 (레거시)")
    print("=" * 60)

    if not GMAIL_ADDRESS or not GMAIL_APP_PASSWORD:
        print("GMAIL_ADDRESS, GMAIL_APP_PASSWORD 환경 변수를 설정하세요.")
        return
    if not SENDER_EMAIL or not RECEIVER_EMAIL:
        print("SENDER_EMAIL, RECEIVER_EMAIL 환경 변수를 설정하세요.")
        return

    print(f"\n[1] 엑셀 '{EXCEL_FILE}' 읽는 중...")
    try:
        headers, customers = read_excel_data(EXCEL_FILE)
        print(f"✓ 총 {len(customers)}명")
    except FileNotFoundError:
        print(f"파일 없음: {EXCEL_FILE}")
        return
    except Exception as e:
        print(f"엑셀 오류: {e}")
        return

    customers_to_send = customers[:MAX_EMAILS]

    print(f"\n[2] SMTP 연결...")
    try:
        smtp = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        smtp.starttls()
        smtp.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        print("✓ 연결됨")
    except Exception as e:
        print(f"SMTP 오류: {e}")
        return

    success_count = 0
    fail_count = 0

    for i, customer in enumerate(customers_to_send, 1):
        try:
            subject = customer.get("메일제목", "안내")
            body = customer.get("메일내용", "")
            customer_name = customer.get("고객명", "고객")
            subject = f"[테스트] {subject}"
            message = create_email_message(SENDER_EMAIL, RECEIVER_EMAIL, subject, body)
            send_email(smtp, GMAIL_ADDRESS, RECEIVER_EMAIL, message)
            success_count += 1
            print(f"  ✅ [{i}/{len(customers_to_send)}] {customer_name}")
            if i < len(customers_to_send):
                time.sleep(1)
        except Exception as e:
            fail_count += 1
            print(f"  ❌ [{i}/{len(customers_to_send)}] {e}")

    smtp.quit()
    print(f"\n완료: 성공 {success_count}, 실패 {fail_count}")


if __name__ == "__main__":
    main()
