"""
엑셀 파일의 고객에게 이메일 발송 프로그램
Gmail SMTP를 사용하여 이메일을 발송합니다.

엑셀 파일 구조:
- 고객번호, 고객명, 주문상품, 수량, 발송일, 운송장번호, 메일제목, 메일내용
"""

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook
import time

# ============== 설정 ==============
EXCEL_FILE = "축산메일실습용.xlsx"
SENDER_EMAIL = "apingdola@naver.com"  # 보내는 사람 이메일 (표시용)
RECEIVER_EMAIL = "apingdola@naver.com"  # 테스트용 - 받는 사람 이메일 (고정)
MAX_EMAILS = 10  # 최대 발송 개수

# Gmail SMTP 설정
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

# Gmail 계정 정보 (앱 비밀번호 사용 필요)
# Gmail 앱 비밀번호 생성 방법:
# 1. Google 계정 설정 -> 보안 -> 2단계 인증 활성화
# 2. Google 계정 설정 -> 보안 -> 앱 비밀번호 생성
# 3. 앱 이름 입력 후 생성된 16자리 비밀번호 사용
GMAIL_ADDRESS = "apingdola@gmail.com"  # Gmail 주소
GMAIL_APP_PASSWORD = "clfneafdmyczijwy"  # Gmail 앱 비밀번호


def read_excel_data(file_path):
    """엑셀 파일에서 고객 데이터 읽기"""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    
    # 첫 번째 행은 헤더로 가정
    headers = [cell.value for cell in sheet[1]]
    print(f"엑셀 파일 컬럼: {headers}")
    
    customers = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if any(row):  # 빈 행이 아닌 경우
            customer = dict(zip(headers, row))
            customers.append(customer)
    
    return headers, customers


def create_email_message(sender, receiver, subject, body):
    """이메일 메시지 생성"""
    message = MIMEMultipart()
    message["From"] = sender
    message["To"] = receiver
    message["Subject"] = subject
    message.attach(MIMEText(body, "plain", "utf-8"))
    return message


def send_email(smtp_connection, sender, receiver, message):
    """이메일 발송"""
    smtp_connection.sendmail(sender, receiver, message.as_string())


def main():
    print("=" * 60)
    print("📧 축산 고객 이메일 발송 프로그램")
    print("=" * 60)
    
    # 1. 엑셀 파일 읽기
    print(f"\n[1] 엑셀 파일 '{EXCEL_FILE}' 읽는 중...")
    try:
        headers, customers = read_excel_data(EXCEL_FILE)
        print(f"✓ 총 {len(customers)}명의 고객 데이터를 읽었습니다.")
    except FileNotFoundError:
        print(f"❌ 오류: '{EXCEL_FILE}' 파일을 찾을 수 없습니다.")
        return
    except Exception as e:
        print(f"❌ 엑셀 파일 읽기 오류: {e}")
        return
    
    # 상위 10개만 선택
    customers_to_send = customers[:MAX_EMAILS]
    print(f"✓ 발송 대상: {len(customers_to_send)}명 (최대 {MAX_EMAILS}명)")
    
    # 고객 데이터 미리보기
    print("\n[고객 목록 미리보기]")
    print("-" * 60)
    for i, customer in enumerate(customers_to_send, 1):
        customer_name = customer.get('고객명', 'N/A')
        product = customer.get('주문상품', 'N/A')
        subject = customer.get('메일제목', 'N/A')
        print(f"  {i}. {customer_name} | {product} | {subject[:30]}...")
    print("-" * 60)
    
    # 2. SMTP 연결
    print(f"\n[2] Gmail SMTP 서버 연결 중...")
    
    if GMAIL_ADDRESS == "YOUR_GMAIL_ADDRESS@gmail.com" or GMAIL_APP_PASSWORD == "YOUR_APP_PASSWORD":
        print("\n" + "!" * 60)
        print("⚠️  Gmail 계정 정보가 설정되지 않았습니다!")
        print("")
        print("send_email.py 파일에서 다음을 수정하세요:")
        print("  - GMAIL_ADDRESS: 본인의 Gmail 주소")
        print("  - GMAIL_APP_PASSWORD: Gmail 앱 비밀번호 (16자리)")
        print("")
        print("📋 Gmail 앱 비밀번호 생성 방법:")
        print("  1. https://myaccount.google.com 접속")
        print("  2. 보안 → 2단계 인증 활성화")
        print("  3. 보안 → 앱 비밀번호 클릭")
        print("  4. 앱 이름 입력 후 '만들기' 클릭")
        print("  5. 생성된 16자리 비밀번호 복사 (공백 제거 후 입력)")
        print("!" * 60)
        return
    
    try:
        smtp = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        smtp.starttls()  # TLS 보안 연결
        smtp.login(GMAIL_ADDRESS, GMAIL_APP_PASSWORD)
        print(f"✓ SMTP 서버 연결 성공! ({SMTP_SERVER}:{SMTP_PORT})")
    except smtplib.SMTPAuthenticationError:
        print("❌ Gmail 인증 실패. 이메일 주소와 앱 비밀번호를 확인하세요.")
        return
    except Exception as e:
        print(f"❌ SMTP 연결 오류: {e}")
        return
    
    # 3. 이메일 발송
    print(f"\n[3] 이메일 발송 시작")
    print(f"    📬 테스트 모드: 모든 메일을 {RECEIVER_EMAIL}로 발송")
    print("")
    
    success_count = 0
    fail_count = 0
    
    for i, customer in enumerate(customers_to_send, 1):
        try:
            # 엑셀 파일의 메일제목과 메일내용 사용
            subject = customer.get('메일제목', f'[팜앤푸드] 발송 안내')
            body = customer.get('메일내용', '안녕하세요, 발송 안내 메일입니다.')
            customer_name = customer.get('고객명', '고객')
            
            # 테스트임을 표시하기 위해 제목에 [테스트] 추가
            subject = f"[테스트] {subject}"
            
            # 이메일 메시지 생성
            message = create_email_message(SENDER_EMAIL, RECEIVER_EMAIL, subject, body)
            
            # 이메일 발송
            send_email(smtp, GMAIL_ADDRESS, RECEIVER_EMAIL, message)
            
            success_count += 1
            print(f"  ✅ [{i:2}/{len(customers_to_send)}] {customer_name}님 - 발송 성공")
            
            # 연속 발송 시 딜레이 (스팸 방지)
            if i < len(customers_to_send):
                time.sleep(1)
                
        except Exception as e:
            fail_count += 1
            print(f"  ❌ [{i:2}/{len(customers_to_send)}] {customer_name}님 - 발송 실패: {e}")
    
    # 4. 연결 종료
    smtp.quit()
    
    # 5. 결과 출력
    print("\n" + "=" * 60)
    print("📊 이메일 발송 완료!")
    print(f"    ✅ 성공: {success_count}건")
    print(f"    ❌ 실패: {fail_count}건")
    print(f"    📥 수신함: {RECEIVER_EMAIL}")
    print("=" * 60)


if __name__ == "__main__":
    main()

